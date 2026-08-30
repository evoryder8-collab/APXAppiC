"""Source-specific adapters for the APEX Food Knowledge Corpus.

Each adapter yields the same canonical record shape. Values are transported as
publisher evidence: special markers remain statuses and source units remain
attached to the original value.
"""

from __future__ import annotations

import csv
import gzip
import io
import json
from pathlib import Path
import re
import subprocess
import tempfile
from typing import Any, Callable, Iterable, Iterator, Optional
import unicodedata
import uuid
import zipfile


CORPUS_NAMESPACE = uuid.UUID("cf87d423-ad7c-506f-928b-1867ac050761")

TRACE_MARKERS = {"tr", "tr.", "trace", "traces"}
NOT_MEASURED_MARKERS = {
    "n",
    "nd",
    "n.d",
    "n.d.",
    "n/a",
    "na",
    "not measured",
    "not analysed",
    "not analyzed",
    "not determined",
}
MISSING_MARKERS = {"", "-", "—", "..", "...", "null", "none", "*"}

SOURCE_PRIORITIES = {
    "swiss-fsvo": 10,
    "usda-foundation": 20,
    "uk-cofid": 22,
    "fr-ciqual": 23,
    "dk-frida": 24,
    "jp-mext": 25,
    "au-afcd": 26,
    "se-livsmedelsverket": 27,
    "ca-cnf": 28,
    "no-matvaretabellen": 29,
    "fi-fineli": 30,
    "nz-concise": 31,
    "usda-fndds": 35,
    "usda-sr-legacy": 40,
    "usda-branded": 60,
}


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalized_search_text(value: Any) -> str:
    decomposed = unicodedata.normalize("NFKD", normalize_text(value)).casefold()
    without_marks = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", " ", without_marks).strip()


def deterministic_record_id(source_key: str, source_record_id: str) -> str:
    return str(uuid.uuid5(CORPUS_NAMESPACE, f"{source_key}:{source_record_id}"))


def parse_value(raw_value: Any) -> tuple[str, float | None, str]:
    original = "" if raw_value is None else str(raw_value).strip()
    folded = original.casefold()
    if folded in MISSING_MARKERS:
        return "missing", None, original
    if folded in TRACE_MARKERS:
        return "trace", None, original
    if folded in NOT_MEASURED_MARKERS:
        return "not_measured", None, original
    if original.startswith("<"):
        return "below_detection", None, original

    parenthetical = original.startswith("(") and original.endswith(")")
    numeric = original.strip("() ").replace("\u00a0", "").replace(" ", "")
    if numeric.count(",") == 1 and "." not in numeric:
        numeric = numeric.replace(",", ".")
    try:
        return ("estimated" if parenthetical else "measured"), float(numeric), original
    except ValueError as error:
        raise ValueError(f"unsupported nutrient observation: {original!r}") from error


def evidence(
    source_code: Any,
    name: Any,
    unit: Any,
    raw_value: Any,
    *,
    method: Any = None,
    reference: Any = None,
    forced_status: str | None = None,
) -> dict[str, Any]:
    status, value, original = parse_value(raw_value)
    method_text = normalize_text(method)
    if forced_status in {"trace", "below_detection", "not_measured", "missing"}:
        status, value = forced_status, None
    elif forced_status in {"measured", "calculated", "estimated"}:
        status = forced_status
    elif status == "measured":
        folded_method = method_text.casefold()
        if "calculat" in folded_method or "summation" in folded_method:
            status = "calculated"
        elif "estimat" in folded_method or "borrowed" in folded_method or "imputed" in folded_method:
            status = "estimated"

    source_code_text = normalize_text(source_code) or normalize_text(name)
    return {
        "nutrient_code": canonical_nutrient_code(source_code_text, name, unit),
        "source_nutrient_code": source_code_text,
        "original_nutrient_name": normalize_text(name) or source_code_text,
        "unit": normalize_text(unit) or "unknown",
        "value": value,
        "observation_status": status,
        "original_value_text": original,
        "derivation_method": method_text or None,
        "source_reference": normalize_text(reference) or None,
    }


def canonical_nutrient_code(source_code: Any, name: Any, unit: Any) -> str:
    code = normalize_text(source_code).upper().replace(" ", "_")
    label = normalize_text(name).casefold()
    unit_text = normalize_text(unit).casefold()
    if "energy" in label and "kcal" in unit_text:
        return "ENERC_KCAL"
    if code in {"KCALS", "ENERC_KCAL", "208", "1008"}:
        return "ENERC_KCAL"
    if code in {"PROT", "PROT-", "203", "1003"} or label == "protein":
        return "PROT"
    if code in {"FAT", "FAT-", "204", "1004"} or label in {"fat", "total fat", "fat, total"}:
        return "FAT"
    if code in {"CHO", "CHOAVL", "CHOAVLDF-", "205", "1005"} or label in {
        "carbohydrate",
        "carbohydrates, available",
        "available carbohydrates",
        "carbohydrate, by difference",
    }:
        return "CHOAVL"
    if code in {"FIBT", "AOACFIB", "FIB-", "291", "1079"} or "dietary fibre" in label or "dietary fiber" in label:
        return "FIBT"
    if code in {"SUGAR", "TOTSUG", "SUGARS", "269", "2000"} or label in {"sugars", "sugars, total", "sugar total", "total sugars"}:
        return "SUGAR"
    if code in {"FASAT", "SATFOD", "606", "1258"} or label in {
        "fatty acids, saturated, total",
        "fatty acids, total saturated",
    }:
        return "FASAT"
    if code in {"NACL", "SALT"} or label in {"salt", "salt, nacl", "salt labelling"}:
        return "NACL"
    if code in {"WATER", "255", "1051"} or label in {"water", "moisture"}:
        return "WATER"
    return code or "UNMAPPED"


def make_record(
    source: dict[str, Any],
    source_record_id: Any,
    name: Any,
    nutrients: list[dict[str, Any]],
    *,
    basis_kind: str = "per_100g",
    basis_amount: float = 100,
    basis_unit: str = "g",
    aliases: Iterable[Any] = (),
    language: str = "en",
    brand: Any = None,
    barcode: Any = None,
    market: Any = None,
    scientific_name: Any = None,
    preparation_state: Any = None,
    edible_portion_percent: Any = None,
    density_g_ml: Any = None,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    source_key = source["key"]
    record_key = normalize_text(source_record_id)
    canonical_name = normalize_text(name)
    if not record_key or not canonical_name:
        raise ValueError(f"{source_key}: record id and name are required")
    alias_values = []
    for alias in aliases:
        cleaned = normalize_text(alias)
        if cleaned and cleaned.casefold() != canonical_name.casefold() and cleaned not in alias_values:
            alias_values.append(cleaned)
    return {
        "id": deterministic_record_id(source_key, record_key),
        "source_key": source_key,
        "source_record_id": record_key,
        "canonical_name": canonical_name,
        "scientific_name": normalize_text(scientific_name) or None,
        "brand": normalize_text(brand) or None,
        "barcode": normalize_text(barcode) or None,
        "market": normalize_text(market) or None,
        "primary_language": language or "und",
        "basis_kind": basis_kind,
        "basis_amount": basis_amount,
        "basis_unit": basis_unit,
        "preparation_state": normalize_text(preparation_state) or None,
        "edible_portion_percent": float(edible_portion_percent) if edible_portion_percent not in (None, "") else None,
        "density_g_ml": float(density_g_ml) if density_g_ml not in (None, "") else None,
        "source_priority": SOURCE_PRIORITIES.get(source_key, 100),
        "source_metadata": metadata or {},
        "aliases": alias_values,
        "nutrients": nutrients,
    }


def _unit_from_heading(heading: Any, default: str = "unknown") -> str:
    text = normalize_text(heading)
    matches = re.findall(r"\(([^()]*)\)", text)
    if not matches:
        return default
    unit = matches[-1].replace("/100g", "").replace("/100 g", "").strip()
    return unit or default


def _limit(records: Iterable[dict[str, Any]], max_records: int | None) -> Iterator[dict[str, Any]]:
    for index, record in enumerate(records):
        if max_records is not None and index >= max_records:
            break
        yield record


def _zip_csv(archive: zipfile.ZipFile, suffix: str) -> Iterator[dict[str, str]]:
    member = next((name for name in archive.namelist() if name.endswith(suffix)), None)
    if member is None:
        raise ValueError(f"archive is missing {suffix}")
    with archive.open(member) as binary:
        with io.TextIOWrapper(binary, encoding="utf-8-sig", newline="") as text:
            yield from csv.DictReader(text)


def parse_usda(
    artifact: Path,
    source: dict[str, Any],
    max_records: int | None,
    brands: list[str],
) -> Iterable[dict[str, Any]]:
    with zipfile.ZipFile(artifact) as archive:
        nutrient_lookup = {row["id"]: row for row in _zip_csv(archive, "/nutrient.csv")}
        branded_meta: dict[str, dict[str, str]] = {}
        is_branded = source["key"] == "usda-branded"
        brand_terms = [term.casefold() for term in brands if normalize_text(term)]
        if is_branded:
            if not brand_terms and max_records is None:
                raise ValueError("USDA Branded requires --brand or --max-records to keep staging bounded")
            for row in _zip_csv(archive, "/branded_food.csv"):
                haystack = " ".join(
                    row.get(field, "")
                    for field in ("brand_owner", "brand_name", "subbrand_name", "short_description")
                ).casefold()
                if brand_terms and not any(term in haystack for term in brand_terms):
                    continue
                branded_meta[row["fdc_id"]] = row
                if max_records is not None and len(branded_meta) >= max_records:
                    break

        foods: dict[str, dict[str, str]] = {}
        for row in _zip_csv(archive, "/food.csv"):
            fdc_id = row["fdc_id"]
            if is_branded and fdc_id not in branded_meta:
                continue
            foods[fdc_id] = row
            if not is_branded and max_records is not None and len(foods) >= max_records:
                break

        nutrients_by_food: dict[str, list[dict[str, Any]]] = {fdc_id: [] for fdc_id in foods}
        for row in _zip_csv(archive, "/food_nutrient.csv"):
            fdc_id = row["fdc_id"]
            if fdc_id not in foods:
                continue
            nutrient = nutrient_lookup.get(row["nutrient_id"])
            if not nutrient or row.get("amount", "") == "":
                continue
            nutrients_by_food[fdc_id].append(
                evidence(
                    nutrient["id"],
                    nutrient["name"],
                    nutrient["unit_name"],
                    row["amount"],
                    reference=f"food_nutrient:{row['id']}",
                )
            )

        def sort_key(value: str) -> tuple[int, str]:
            return (int(value), value) if value.isdigit() else (0, value)

        for fdc_id in sorted(foods, key=sort_key):
            food = foods[fdc_id]
            branded = branded_meta.get(fdc_id, {})
            aliases = [branded.get("short_description")]
            yield make_record(
                source,
                fdc_id,
                food.get("description"),
                nutrients_by_food[fdc_id],
                aliases=aliases,
                brand=branded.get("brand_name") or branded.get("brand_owner"),
                barcode=branded.get("gtin_upc"),
                market=branded.get("market_country") or food.get("market_country"),
                preparation_state=branded.get("preparation_state_code"),
                metadata={
                    "data_type": food.get("data_type"),
                    "publication_date": food.get("publication_date"),
                    "serving_size": branded.get("serving_size"),
                    "serving_size_unit": branded.get("serving_size_unit"),
                    "household_serving": branded.get("household_serving_fulltext"),
                },
            )


def _open_workbook(path: Path):
    try:
        import openpyxl
    except ImportError as error:
        raise ValueError("openpyxl is required for spreadsheet sources") from error
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def parse_swiss(
    artifact: Path,
    source: dict[str, Any],
    max_records: int | None,
    brands: list[str],
) -> Iterable[dict[str, Any]]:
    workbook = _open_workbook(artifact)

    def records() -> Iterator[dict[str, Any]]:
        for sheet_name in ("Generic Foods", "Branded foods"):
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            next(rows, None)
            next(rows, None)
            headers = list(next(rows))
            for row in rows:
                if not row or row[0] in (None, "") or row[3] in (None, ""):
                    continue
                matrix = normalize_text(row[7]).casefold()
                basis_kind = "per_100ml" if "100ml" in matrix else "per_100g"
                basis_unit = "ml" if basis_kind == "per_100ml" else "g"
                nutrients: list[dict[str, Any]] = []
                for index in range(8, len(headers), 3):
                    heading = headers[index]
                    if not heading:
                        continue
                    unit = _unit_from_heading(heading)
                    if unit == "unknown":
                        continue
                    method = row[index + 1] if index + 1 < len(row) else None
                    reference = row[index + 2] if index + 2 < len(row) else None
                    nutrients.append(
                        evidence(
                            heading,
                            re.sub(r"\s*\([^)]*\)\s*$", "", normalize_text(heading)),
                            unit,
                            row[index] if index < len(row) else None,
                            method=method,
                            reference=reference,
                        )
                    )
                aliases = re.split(r"[;|]", normalize_text(row[4])) if row[4] else []
                yield make_record(
                    source,
                    f"{sheet_name.casefold().replace(' ', '-')}:{row[0]}",
                    row[3],
                    nutrients,
                    basis_kind=basis_kind,
                    basis_unit=basis_unit,
                    aliases=aliases,
                    brand="Swiss branded foods" if sheet_name == "Branded foods" else None,
                    density_g_ml=row[6],
                    metadata={"category": row[5], "matrix_unit": row[7]},
                )

    return _limit(records(), max_records)


def parse_cofid(
    artifact: Path,
    source: dict[str, Any],
    max_records: int | None,
    brands: list[str],
) -> Iterable[dict[str, Any]]:
    workbook = _open_workbook(artifact)
    sheet = workbook["1.3 Proximates"]
    rows = sheet.iter_rows(values_only=True)
    headings = list(next(rows))
    codes = list(next(rows))
    names = list(next(rows))

    def records() -> Iterator[dict[str, Any]]:
        for row in rows:
            if not row or not row[0] or not row[1]:
                continue
            nutrients = [
                evidence(
                    codes[index] or headings[index],
                    names[index] or headings[index],
                    _unit_from_heading(headings[index]),
                    row[index] if index < len(row) else None,
                    reference=row[5],
                )
                for index in range(7, len(headings))
                if headings[index]
            ]
            yield make_record(
                source,
                row[0],
                row[1],
                nutrients,
                preparation_state=row[2],
                metadata={"group": row[3], "description": row[2], "references": row[5]},
            )

    return _limit(records(), max_records)


def _converted_xlsx(artifact: Path) -> tuple[tempfile.TemporaryDirectory[str] | None, Path]:
    if artifact.suffix.casefold() == ".xlsx":
        return None, artifact
    temporary = tempfile.TemporaryDirectory(prefix="apex-food-corpus-")
    result = subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", temporary.name, str(artifact)],
        check=False,
        capture_output=True,
        text=True,
    )
    converted = Path(temporary.name) / f"{artifact.stem}.xlsx"
    if result.returncode != 0 or not converted.is_file():
        temporary.cleanup()
        raise ValueError(f"could not convert {artifact.name}: {result.stderr.strip()}")
    return temporary, converted


def parse_ciqual(
    artifact: Path,
    source: dict[str, Any],
    max_records: int | None,
    brands: list[str],
) -> Iterable[dict[str, Any]]:
    temporary, converted = _converted_xlsx(artifact)
    workbook = _open_workbook(converted)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headings = list(next(rows))

    def records() -> Iterator[dict[str, Any]]:
        try:
            for row in rows:
                if not row or not row[6] or not row[7]:
                    continue
                nutrients = [
                    evidence(
                        headings[index],
                        re.sub(r"\s*\([^)]*/100g\)\s*$", "", normalize_text(headings[index])),
                        _unit_from_heading(headings[index]),
                        row[index] if index < len(row) else None,
                    )
                    for index in range(9, len(headings))
                    if headings[index]
                ]
                yield make_record(
                    source,
                    row[6],
                    row[7],
                    nutrients,
                    scientific_name=row[8],
                    metadata={
                        "group": row[3],
                        "subgroup": row[4],
                        "sub_subgroup": row[5],
                    },
                )
        finally:
            if temporary is not None:
                temporary.cleanup()

    return _limit(records(), max_records)


def parse_frida(
    artifact: Path,
    source: dict[str, Any],
    max_records: int | None,
    brands: list[str],
) -> Iterable[dict[str, Any]]:
    workbook = _open_workbook(artifact)
    sheet = workbook["Data_Table"]
    rows = sheet.iter_rows(values_only=True)
    next(rows)
    names = list(next(rows))
    units = list(next(rows))
    ids = list(next(rows))

    def records() -> Iterator[dict[str, Any]]:
        for row in rows:
            if not row or not row[2] or not row[1]:
                continue
            nutrients = [
                evidence(
                    ids[index],
                    names[index],
                    units[index],
                    row[index] if index < len(row) else None,
                )
                for index in range(3, len(names))
                if ids[index] not in (None, "") and names[index]
            ]
            yield make_record(source, row[2], row[1], nutrients, aliases=[row[0]])

    return _limit(records(), max_records)


def parse_mext(
    artifact: Path,
    source: dict[str, Any],
    max_records: int | None,
    brands: list[str],
) -> Iterable[dict[str, Any]]:
    workbook = _open_workbook(artifact)
    sheet = workbook["表全体"]
    rows = sheet.iter_rows(values_only=True)
    header_rows = [list(next(rows)) for _ in range(12)]
    units = header_rows[10]
    codes = header_rows[11]

    def records() -> Iterator[dict[str, Any]]:
        for row in rows:
            if not row or not row[1] or not row[3]:
                continue
            nutrients = []
            for index in range(4, len(codes)):
                if not codes[index]:
                    continue
                raw = row[index] if index < len(row) else None
                unit = normalize_text(units[index]).strip("()") or "unknown"
                nutrients.append(evidence(codes[index], codes[index], unit, raw))
            yield make_record(
                source,
                row[1],
                row[3],
                nutrients,
                language="ja",
                edible_portion_percent=(100 - float(row[4])) if row[4] not in (None, "", "-") else None,
                metadata={"food_group": row[0], "index_number": row[2]},
            )

    return _limit(records(), max_records)


def parse_afcd(
    artifact: Path,
    source: dict[str, Any],
    max_records: int | None,
    brands: list[str],
) -> Iterable[dict[str, Any]]:
    workbook = _open_workbook(artifact)

    def records() -> Iterator[dict[str, Any]]:
        for sheet_name, basis_kind, basis_unit in (
            ("All solids & liquids per 100 g", "per_100g", "g"),
            ("Liquids only per 100 mL", "per_100ml", "ml"),
        ):
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            next(rows, None)
            next(rows, None)
            headings = list(next(rows))
            for row in rows:
                if not row or not row[0] or not row[3]:
                    continue
                nutrients = [
                    evidence(
                        headings[index],
                        re.sub(r"\s*\([^)]*\)\s*$", "", normalize_text(headings[index])),
                        _unit_from_heading(headings[index]),
                        row[index] if index < len(row) else None,
                        method=row[2],
                    )
                    for index in range(4, len(headings))
                    if headings[index]
                ]
                yield make_record(
                    source,
                    f"{row[0]}:{basis_kind}",
                    row[3],
                    nutrients,
                    basis_kind=basis_kind,
                    basis_unit=basis_unit,
                    metadata={"classification": row[1], "derivation": row[2]},
                )

    return _limit(records(), max_records)


def parse_nz_concise(
    artifact: Path,
    source: dict[str, Any],
    max_records: int | None,
    brands: list[str],
) -> Iterable[dict[str, Any]]:
    workbook = _open_workbook(artifact)
    sheet = workbook["Concisen Tables 14th Edition wi"]
    rows = sheet.iter_rows(values_only=True)
    headings = list(next(rows))
    next(rows, None)
    units = list(next(rows))
    next(rows, None)

    def records() -> Iterator[dict[str, Any]]:
        for row in rows:
            if not row or not row[0] or not row[1] or row[2] != 100:
                continue
            nutrients = [
                evidence(headings[index], headings[index], units[index], row[index] if index < len(row) else None)
                for index in range(3, len(headings))
                if headings[index]
            ]
            yield make_record(source, row[0], row[1], nutrients)

    return _limit(records(), max_records)


def parse_livsmedelsverket(
    artifact: Path,
    source: dict[str, Any],
    max_records: int | None,
    brands: list[str],
) -> Iterable[dict[str, Any]]:
    with artifact.open(encoding="utf-8") as handle:
        foods = {str(food["nummer"]): food for food in json.load(handle)}
    nutrient_path = artifact.with_name("livsmedelsverket_nutrients_en.json.gz")
    with gzip.open(nutrient_path, "rt", encoding="utf-8") as handle:
        nutrient_groups = json.load(handle)

    def records() -> Iterator[dict[str, Any]]:
        for group in nutrient_groups:
            key = str(group["nummer"])
            food = foods.get(key, {})
            nutrients = []
            for item in group.get("naringsvarden", []):
                status_code = normalize_text(item.get("vardetypkod")).upper()
                forced = {
                    "BL": "below_detection",
                    "BE": "estimated",
                    "LZ": "estimated",
                    "MN": "measured",
                }.get(status_code)
                nutrients.append(
                    evidence(
                        item.get("euroFIRkod") or item.get("forkortning"),
                        item.get("namn"),
                        item.get("enhet"),
                        item.get("varde"),
                        method=item.get("metodtyp") or item.get("vardetyp"),
                        reference=item.get("publikation"),
                        forced_status=forced,
                    )
                )
            yield make_record(
                source,
                key,
                food.get("namn") or group.get("namn"),
                nutrients,
                scientific_name=food.get("vetenskapligtNamn"),
                metadata={"food_type": food.get("livsmedelsTyp"), "project": food.get("projekt")},
            )

    return _limit(records(), max_records)


def parse_canadian_nutrient_file(
    artifact: Path,
    source: dict[str, Any],
    max_records: int | None,
    brands: list[str],
) -> Iterable[dict[str, Any]]:
    with artifact.open(encoding="utf-8") as handle:
        foods = {str(food["food_code"]): food for food in json.load(handle)}
    names_path = artifact.with_name("cnf_nutrientname_en.json")
    with names_path.open(encoding="utf-8") as handle:
        names = {str(item["nutrient_name_id"]): item for item in json.load(handle)}
    amounts_path = artifact.with_name("cnf_nutrient_amounts_en.json.gz")
    with gzip.open(amounts_path, "rt", encoding="utf-8") as handle:
        amount_groups = json.load(handle)

    def records() -> Iterator[dict[str, Any]]:
        for group in amount_groups:
            key = str(group["food_code"])
            food = foods.get(key, {})
            nutrients = []
            for item in group.get("nutrients", []):
                nutrient_name = names.get(str(item["nutrient_name_id"]), {})
                nutrients.append(
                    evidence(
                        nutrient_name.get("tagname") or nutrient_name.get("nutrient_symbol") or item["nutrient_name_id"],
                        nutrient_name.get("nutrient_web_name") or item.get("nutrient_web_name"),
                        nutrient_name.get("unit"),
                        item.get("nutrient_value"),
                        reference=f"CNF-source:{item.get('nutrient_source_id')}",
                    )
                )
            yield make_record(
                source,
                key,
                food.get("food_description") or group.get("food_description"),
                nutrients,
                market="Canada",
            )

    return _limit(records(), max_records)


def parse_matvaretabellen(
    artifact: Path,
    source: dict[str, Any],
    max_records: int | None,
    brands: list[str],
) -> Iterable[dict[str, Any]]:
    with artifact.open(encoding="utf-8") as handle:
        payload = json.load(handle)

    def records() -> Iterator[dict[str, Any]]:
        for food in payload.get("foods", []):
            nutrients = []
            calories = food.get("calories") or {}
            if calories.get("quantity") is not None:
                nutrients.append(
                    evidence("ENERC_KCAL", "Energy", calories.get("unit", "kcal"), calories["quantity"], reference=calories.get("sourceId"))
                )
            for item in food.get("constituents", []):
                nutrients.append(
                    evidence(
                        item.get("nutrientId"),
                        item.get("nutrientId"),
                        item.get("unit"),
                        item.get("quantity"),
                        reference=item.get("sourceId"),
                    )
                )
            edible = (food.get("ediblePart") or {}).get("percent")
            yield make_record(
                source,
                food.get("foodId"),
                food.get("foodName"),
                nutrients,
                aliases=food.get("searchKeywords") or [],
                scientific_name=food.get("latinName"),
                market="Norway",
                edible_portion_percent=edible,
                metadata={"uri": food.get("uri"), "food_group_id": food.get("foodGroupId")},
            )

    return _limit(records(), max_records)


def _read_zip_semicolon(archive: zipfile.ZipFile, member: str) -> list[dict[str, str]]:
    with archive.open(member) as binary:
        raw = binary.read()
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError(f"could not decode {member}")
    return list(csv.DictReader(io.StringIO(text), delimiter=";"))


def parse_fineli(
    artifact: Path,
    source: dict[str, Any],
    max_records: int | None,
    brands: list[str],
) -> Iterable[dict[str, Any]]:
    with zipfile.ZipFile(artifact) as archive:
        foods = {row["FOODID"]: row for row in _read_zip_semicolon(archive, "food.csv")}
        english = {row["FOODID"]: row["FOODNAME"] for row in _read_zip_semicolon(archive, "foodname_EN.csv")}
        finnish = {row["FOODID"]: row["FOODNAME"] for row in _read_zip_semicolon(archive, "foodname_FI.csv")}
        swedish = {row["FOODID"]: row["FOODNAME"] for row in _read_zip_semicolon(archive, "foodname_SV.csv")}
        components = {row["EUFDNAME"]: row for row in _read_zip_semicolon(archive, "component.csv")}
        values_by_food: dict[str, list[dict[str, str]]] = {}
        for row in _read_zip_semicolon(archive, "component_value.csv"):
            values_by_food.setdefault(row["FOODID"], []).append(row)

    def records() -> Iterator[dict[str, Any]]:
        for food_id in sorted(foods, key=lambda value: int(value)):
            name = english.get(food_id) or foods[food_id].get("FOODNAME")
            if not name:
                continue
            nutrients = []
            for row in values_by_food.get(food_id, []):
                component = components.get(row["EUFDNAME"], {})
                nutrients.append(
                    evidence(
                        row["EUFDNAME"],
                        row["EUFDNAME"],
                        component.get("COMPUNIT"),
                        row.get("BESTLOC"),
                        method=f"{row.get('ACQTYPE', '')}:{row.get('METHTYPE', '')}:{row.get('METHIND', '')}",
                    )
                )
            yield make_record(
                source,
                food_id,
                name,
                nutrients,
                aliases=[finnish.get(food_id), swedish.get(food_id)],
                market="Finland",
                preparation_state=foods[food_id].get("PROCESS"),
                edible_portion_percent=foods[food_id].get("EDPORT"),
            )

    return _limit(records(), max_records)


Parser = Callable[[Path, dict[str, Any], Optional[int], list[str]], Iterable[dict[str, Any]]]

SUPPORTED_PARSERS: dict[str, Parser] = {
    "usda_fooddata_central_v1": parse_usda,
    "swiss_fsvo_v1": parse_swiss,
    "cofid_v1": parse_cofid,
    "ciqual_v1": parse_ciqual,
    "frida_v1": parse_frida,
    "mext_v1": parse_mext,
    "afcd_v1": parse_afcd,
    "livsmedelsverket_v1": parse_livsmedelsverket,
    "canadian_nutrient_file_v1": parse_canadian_nutrient_file,
    "matvaretabellen_v1": parse_matvaretabellen,
    "fineli_v1": parse_fineli,
    "nz_concise_v1": parse_nz_concise,
}


def parse_source(
    artifact: Path,
    source: dict[str, Any],
    *,
    max_records: int | None,
    brands: list[str],
) -> Iterable[dict[str, Any]]:
    parser_name = source["parser"]
    adapter = SUPPORTED_PARSERS.get(parser_name)
    if adapter is None:
        raise ValueError(f"unsupported parser: {parser_name}")
    return adapter(artifact, source, max_records, brands)


def search_projection(record: dict[str, Any]) -> dict[str, Any]:
    macro_fields = {
        "ENERC_KCAL": "kcal",
        "PROT": "protein_g",
        "CHOAVL": "carbs_g",
        "FAT": "fat_g",
        "FIBT": "fibre_g",
        "SUGAR": "sugar_g",
        "FASAT": "saturated_fat_g",
        "NACL": "salt_g",
        "WATER": "water_g",
    }
    macros: dict[str, float | None] = {field: None for field in macro_fields.values()}
    for nutrient in record["nutrients"]:
        field = macro_fields.get(nutrient["nutrient_code"])
        if field and macros[field] is None and nutrient["value"] is not None:
            macros[field] = nutrient["value"]

    search_parts = [
        record["canonical_name"],
        *record["aliases"],
        record.get("brand"),
        record.get("barcode"),
        record.get("market"),
    ]
    return {
        "record_id": record["id"],
        "source_key": record["source_key"],
        "source_record_id": record["source_record_id"],
        "name": record["canonical_name"],
        "names_i18n": {record["primary_language"]: record["canonical_name"]},
        "aliases": record["aliases"],
        "brand": record.get("brand"),
        "barcode": record.get("barcode"),
        "market": record.get("market"),
        "basis_kind": record["basis_kind"],
        "preparation_state": record.get("preparation_state"),
        **macros,
        "source_priority": record["source_priority"],
        "search_text": normalized_search_text(" ".join(str(part or "") for part in search_parts)),
    }
